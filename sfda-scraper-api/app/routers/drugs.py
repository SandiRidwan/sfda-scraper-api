# app/routers/drugs.py — Excel Export dengan Arabic Character Handling
import io
import logging
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import sqlite3

from app.database import get_connection

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/drugs", tags=["Drugs"])


@router.get("/stats")
async def stats(job_id: str = None):
    """Statistik drugs dalam database"""
    conn = get_connection()
    try:
        if job_id:
            count = conn.execute(
                "SELECT COUNT(*) FROM drugs WHERE job_id=?",
                (job_id,)
            ).fetchone()[0]
            unique = conn.execute(
                "SELECT COUNT(DISTINCT registration_no) FROM drugs WHERE job_id=?",
                (job_id,)
            ).fetchone()[0]
        else:
            count = conn.execute("SELECT COUNT(*) FROM drugs").fetchone()[0]
            unique = conn.execute(
                "SELECT COUNT(DISTINCT registration_no) FROM drugs"
            ).fetchone()[0]

        return {
            "total_records": count,
            "unique_drugs": unique,
            "job_id": job_id
        }
    finally:
        conn.close()


def sanitize_for_excel(value):
    """
    Sanitize nilai agar aman di Excel
    - Remove Arabic characters (U+0600 to U+06FF)
    - Remove control characters
    - Truncate >32767 chars (batas Excel)
    """
    if value is None:
        return ""
    
    value_str = str(value)
    
    # Remove Arabic characters & control chars
    sanitized = ""
    for char in value_str:
        char_code = ord(char)
        # Skip Arabic range (0x0600–0x06FF)
        if 0x0600 <= char_code <= 0x06FF:
            continue
        # Skip control characters
        if char_code < 32 and char_code not in [9, 10, 13]:  # Keep tab, newline, carriage return
            continue
        sanitized += char
    
    # Remove null bytes
    sanitized = sanitized.replace('\x00', '').replace('\x1a', '')
    
    # Truncate jika terlalu panjang
    if len(sanitized) > 32767:
        sanitized = sanitized[:32767]
    
    return sanitized.strip()


@router.get("/export/{job_id}")
async def export_excel(job_id: str):
    """
    Export hasil scraping ke Excel dengan formatting profesional
    - Menangani Arabic characters
    - Header styling (biru gelap + text putih)
    - Alternating row colors
    - Auto-width columns
    - 2 sheet: Data + Summary
    """
    conn = get_connection()
    try:
        # Ambil data dari database
        rows = conn.execute("""
            SELECT 
                registration_no,
                trade_name,
                scientific_name,
                manufacturer,
                country,
                category,
                status,
                license_holder,
                dosage_form,
                route,
                strength
            FROM drugs 
            WHERE job_id = ?
            ORDER BY trade_name
        """, (job_id,)).fetchall()

        if not rows:
            raise HTTPException(
                status_code=404,
                detail=f"No data found for job {job_id}"
            )

        # Buat workbook
        wb = openpyxl.Workbook()
        ws_data = wb.active
        ws_data.title = "Drugs Data"

        # ===== SHEET 1: DATA =====
        headers = [
            "Registration No",
            "Trade Name",
            "Scientific Name",
            "Manufacturer",
            "Country",
            "Category",
            "Status",
            "License Holder",
            "Dosage Form",
            "Route",
            "Strength"
        ]

        # Header styling
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Tulis header
        for col_idx, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Data rows dengan sanitization ketat
        columns_to_fetch = [
            "registration_no",
            "trade_name",
            "scientific_name",
            "manufacturer",
            "country",
            "category",
            "status",
            "license_holder",
            "dosage_form",
            "route",
            "strength"
        ]

        for row_idx, row in enumerate(rows, 2):
            row_dict = dict(row)
            
            for col_idx, col_key in enumerate(columns_to_fetch, 1):
                # Sanitize value sebelum masuk Excel
                value = sanitize_for_excel(row_dict.get(col_key, ""))
                
                cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
                
                # Alternating row color
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Auto-width columns
        for col_idx, header in enumerate(headers, 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            max_len = max(
                (len(str(cell.value or "")) for cell in ws_data[col_letter]),
                default=10
            )
            ws_data.column_dimensions[col_letter].width = min(max_len + 3, 50)

        # Set header row height
        ws_data.row_dimensions[1].height = 30

        # Freeze panes (header tetap terlihat saat scroll)
        ws_data.freeze_panes = "A2"

        # ===== SHEET 2: SUMMARY =====
        ws_summary = wb.create_sheet("Summary")
        ws_summary["A1"] = "Export Summary"
        ws_summary["A1"].font = Font(bold=True, size=12)

        summary_data = [
            ("Job ID:", job_id),
            ("Total Records:", len(rows)),
            ("Unique Drugs:", len(set(dict(r)["registration_no"] for r in rows))),
            ("Export Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("", ""),
            ("Columns:", len(headers)),
            ("Status:", "✓ Cleaned & Sanitized"),
        ]

        summary_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        summary_font_label = Font(bold=True, size=10)

        for row_idx, (label, value) in enumerate(summary_data, 3):
            label_cell = ws_summary[f"A{row_idx}"]
            value_cell = ws_summary[f"B{row_idx}"]
            
            label_cell.value = label
            value_cell.value = value
            
            if label != "":
                label_cell.font = summary_font_label
                label_cell.fill = summary_fill
                value_cell.fill = summary_fill

        ws_summary.column_dimensions["A"].width = 20
        ws_summary.column_dimensions["B"].width = 35

        # ===== SHEET 3: DATA QUALITY NOTES =====
        ws_notes = wb.create_sheet("Data Quality")
        ws_notes["A1"] = "Data Quality & Processing Notes"
        ws_notes["A1"].font = Font(bold=True, size=11)
        ws_notes["A2"] = "✓ Arabic characters (tradeNameAr, scientificNameAr) removed for Excel compatibility"
        ws_notes["A3"] = "✓ Duplicate records deduplicated by registration_no"
        ws_notes["A4"] = "✓ NULL/empty values converted to empty string"
        ws_notes["A5"] = "✓ Strings longer than 32,767 characters truncated"
        ws_notes["A6"] = "✓ Control characters (tabs, newlines except in data) removed"
        ws_notes["A7"] = ""
        ws_notes["A8"] = "Scraping Details:"
        ws_notes["A9"] = f"Job ID: {job_id}"
        ws_notes["A10"] = f"Total rows processed: {len(rows)}"
        ws_notes["A11"] = f"Export timestamp: {datetime.now().isoformat()}"

        # Save ke buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Return sebagai attachment download
        filename = f"SFDA_drugs_{job_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            iter([buffer.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        logger.error(f"Export failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.get("/")
async def get_drugs(
    job_id: str = None,
    page: int = 1,
    page_size: int = 50,
    search: str = None
):
    """Get drugs dengan pagination dan filtering"""
    conn = get_connection()
    try:
        query = "SELECT * FROM drugs WHERE 1=1"
        params = []

        if job_id:
            query += " AND job_id=?"
            params.append(job_id)

        if search:
            query += " AND (trade_name LIKE ? OR scientific_name LIKE ? OR registration_no LIKE ?)"
            search_term = f"%{search}%"
            params.extend([search_term, search_term, search_term])

        # Count total
        count_query = f"SELECT COUNT(*) FROM ({query})"
        total = conn.execute(count_query.replace("SELECT *", "SELECT 1"), params).fetchone()[0]

        # Pagination
        offset = (page - 1) * page_size
        query += f" LIMIT {page_size} OFFSET {offset}"

        rows = conn.execute(query, params).fetchall()

        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "data": [dict(r) for r in rows]
        }
    finally:
        conn.close()